import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO, TextIOWrapper
from typing import Iterable, NamedTuple, Optional
from zoneinfo import ZoneInfo

import pdfplumber
from openpyxl import load_workbook

from firemerge.model import Account, Money, StatementTransaction


def make_notes(meta: dict[str, str]) -> Optional[str]:
    return "\n".join(f"{k}: {v}" for k, v in meta.items()) if meta else None


class StatementReader:
    def __init__(self, data: BytesIO, account: Account, tz: ZoneInfo):
        self.data = data
        self.account = account
        self.tz = tz

    def _read(self) -> Iterable[StatementTransaction]:
        raise NotImplementedError()

    @classmethod
    def read(
        cls, data: BytesIO, account: Account, tz: ZoneInfo
    ) -> list[StatementTransaction]:
        errors = []
        for reader_class in cls.__subclasses__():
            data.seek(0)
            try:
                return list(reader_class(data, account, tz)._read())
            except Exception as e:
                errors.append(e)
        raise ExceptionGroup("Failed to read statement", errors)


class AvalStatementReader(StatementReader):
    HEADER = [
        "Дата операції",
        "Дата обробки операції",
        "Номер картки",
        "Тип операції",
        "Деталі операції",
        "Сума у валюті рахунку",
        "Сума у валюті операції",
        "Валюта операції",
        "Вихідний залишок (у валюті рахунку)",
    ]

    @staticmethod
    def _translate(s: str) -> str:
        return s.replace("\n", " ")

    @staticmethod
    def _money(s: str) -> Money:
        return Money(s.replace(" ", "").replace(",", "."))

    def _read(self) -> Iterable[StatementTransaction]:
        found = False
        with pdfplumber.open(self.data) as pdf:
            for page in pdf.pages:
                for table in page.find_tables():
                    data = [
                        [self._translate(c) if c else "" for c in row]
                        for row in table.extract()
                    ]
                    if data[0] == self.HEADER:
                        found = True
                        for row in data[1:]:
                            yield StatementTransaction(
                                name=row[4],
                                date=datetime.strptime(
                                    row[0], "%d.%m.%Y %H:%M"
                                ).replace(tzinfo=self.tz),
                                amount=self._money(row[5]),
                                foreign_amount=self._money(row[6])
                                if row[5] == row[6]
                                else None,
                                foreign_currency_code=row[7]
                                if row[5] == row[6]
                                else None,
                                notes=make_notes({"Op Type": row[3], "Description": row[4]}),
                            )
            if not found:
                raise ValueError("Statement not found")


class _ABRecord(NamedTuple):
    account: str
    currency: str
    date: datetime
    doc_number: str
    amount: Optional[Money]
    purpose: str


class AvalBusinessStatementReader(StatementReader):
    HEADER = [
        "ЄДРПОУ",
        "МФО",
        "Рахунок",
        "Валюта",
        "Дата операції",
        "Код операції",
        "МФО банка",
        "Назва банка",
        "Рахунок кореспондента",
        "ЄДРПОУ кореспондента",
        "Кореспондент",
        "Документ",
        "Дата документа",
        "Дебет",
        "Кредит",
        "Призначення платежу",
        "Гривневе покриття",
    ]

    def _read(self) -> Iterable[StatementTransaction]:
        reader = csv.reader(TextIOWrapper(self.data, encoding="cp1251"), delimiter=";")
        header = next(reader)
        if header != self.HEADER:
            print(header)
            raise ValueError("Header mismatch")
        own_records = []
        other_records_map = {}
        for row in reader:
            if row[13] and row[14]:
                raise ValueError("Both debit and credit are present")
            record = _ABRecord(
                account=row[2],
                currency=row[3],
                date=datetime.strptime(row[4], "%d.%m.%Y %H:%M").replace(tzinfo=self.tz),
                doc_number=row[11],
                amount=-Money(row[13]) if row[13] else Money(row[14]),
                purpose=row[15],
            )
            if not record.amount:
                continue
            if record.account == self.account.iban:
                own_records.append(record)
            else:
                other_records_map[record.doc_number] = record
        own_records.sort(key=lambda x: x.date)
        for record in own_records:
            if (other_record := other_records_map.get(record.doc_number)):
                if record.amount * other_record.amount >= 0:
                    raise ValueError("Same direction")
                yield StatementTransaction(
                    name=record.purpose,
                    date=record.date,
                    amount=record.amount,
                    foreign_amount=other_record.amount,
                    foreign_currency_code=other_record.currency,
                    notes=make_notes({"Desciption 1": record.purpose, "Description 2": other_record.purpose, "Account": other_record.account}),
                )
            else:
                yield StatementTransaction(
                    name=record.purpose,
                    date=record.date,
                    amount=record.amount,
                    notes=make_notes({"Description": record.purpose}),
                    foreign_amount=None,
                    foreign_currency_code=None,
                )

class PrivatStatementReader(StatementReader):
    HEADER = [
        "Дата",
        "Категорія",
        "Картка",
        "Опис операції",
        "Сума в валюті картки",
        "Валюта картки",
        "Сума в валюті транзакції",
        "Валюта транзакції",
        "Залишок на кінець періоду",
        "Валюта залишку",
    ]

    @staticmethod
    def _money(s: float | int) -> Money:
        return Money(Decimal(s).quantize(Decimal("0.01")))

    def _read(self) -> Iterable[StatementTransaction]:
        wb = load_workbook(self.data, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("No active sheet")
        seen_header = False
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if not seen_header:
                if values == self.HEADER:
                    seen_header = True
                continue
            assert isinstance(values[4], (float, int))
            assert isinstance(values[6], (float, int))
            yield StatementTransaction(
                name=str(values[3]),
                date=datetime.strptime(str(values[0]), "%d.%m.%Y %H:%M:%S").replace(
                    tzinfo=self.tz
                ),
                amount=self._money(values[4]),
                foreign_amount=self._money(values[6])
                if values[4] == values[6]
                else None,
                foreign_currency_code=str(values[5])
                if values[4] == values[6]
                else None,
                notes=make_notes({"Category": str(values[1]), "Description": str(values[3])}),
            )