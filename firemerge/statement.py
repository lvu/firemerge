from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable
from zoneinfo import ZoneInfo

import pdfplumber
from openpyxl import load_workbook

from .model import Money, StatementTransaction


class StatementReader:

    def __init__(self, data: BytesIO, tz: ZoneInfo):
        self.data = data
        self.tz = tz

    def _read(self) -> Iterable[StatementTransaction]:
        raise NotImplementedError()

    @classmethod
    def read(cls, data: BytesIO, tz: ZoneInfo) -> list[StatementTransaction]:
        errors = []
        for reader_class in cls.__subclasses__():
            try:
                return list(reader_class(data, tz)._read())
            except Exception as e:
                errors.append(e)
        raise ExceptionGroup("Failed to read statement", errors)


class AvalStatementReader(StatementReader):
    HEADER = [
        "Дата операції",
        "Дата обробки операції",
        "Номер картки",
        "Тип операції",
        "Деталі операції",
        "Сума у валюті рахунку",
        "Сума у валюті операції",
        "Валюта операції",
        "Вихідний залишок (у валюті рахунку)",
    ]

    @staticmethod
    def _translate(s: str) -> str:
        return s.replace("\n", " ")

    @staticmethod
    def _money(s: str) -> Money:
        return Money(s.replace(" ", "").replace(",", "."))


    def _read(self) -> Iterable[StatementTransaction]:
        found = False
        with pdfplumber.open(self.data) as pdf:
            for page in pdf.pages:
                for table in page.find_tables():
                    data = [
                        [self._translate(c) if c else "" for c in row] for row in table.extract()
                    ]
                    if data[0] == self.HEADER:
                        found = True
                        for row in data[1:]:
                            yield StatementTransaction(
                                name=row[4],
                                date=datetime.strptime(row[0], "%d.%m.%Y %H:%M").replace(
                                    tzinfo=self.tz
                                ),
                                amount=self._money(row[5]),
                                foreign_amount=self._money(row[6]) if row[5] == row[6] else None,
                                foreign_currency_code=row[7] if row[5] == row[6] else None,
                                meta={"Op Type": row[3], "Description": row[4]},
                            )
            if not found:
                raise ValueError("Statement not found")


class PrivatStatementReader(StatementReader):
    HEADER = [
        "Дата",
        "Категорія",
        "Картка",
        "Опис операції",
        "Сума в валюті картки",
        "Валюта картки",
        "Сума в валюті транзакції",
        "Валюта транзакції",
        "Залишок на кінець періоду",
        "Валюта залишку"
    ]

    @staticmethod
    def _money(s: float | int) -> Money:
        return Money(Decimal(s).quantize(Decimal("0.01")))

    def _read(self) -> Iterable[StatementTransaction]:
        wb = load_workbook(self.data, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("No active sheet")
        seen_header = False
        for row in ws.iter_rows():
            row = [cell.value for cell in row]
            if not seen_header:
                if row == self.HEADER:
                    seen_header = True
                continue
            assert isinstance(row[4], (float, int))
            assert isinstance(row[6], (float, int))
            yield StatementTransaction(
                name=str(row[3]),
                date=datetime.strptime(str(row[0]), "%d.%m.%Y %H:%M:%S").replace(tzinfo=self.tz),
                amount=self._money(row[4]),
                foreign_amount=self._money(row[6]) if row[4] == row[6] else None,
                foreign_currency_code=str(row[5]) if row[4] == row[6] else None,
                meta={"Category": str(row[1]), "Description": str(row[3])},
            )